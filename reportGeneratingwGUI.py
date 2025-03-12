from tkinter import *
from tkinter import ttk

import mysql.connector
import openpyxl

mydb = mysql.connector.connect(
    user='root',  # username
    password='R00t',  # password
    host='192.168.250.10',
    port='3307',
    database='hospital_dbo'
)

mycursor = mydb.cursor()
win= Tk()
win.geometry("500x400")



def display_text():
    global path2
    global column_no2
    global row_start2
    global column_letter2
    global dispo

    selected_dispo = dispo.get()

    path= path2.get()
    column_no = int(column_no2.get())
    row_start = int(row_start2.get())
    column_letter = column_letter2.get()

    wb_obj = openpyxl.load_workbook(path+".xlsx")

    sheet_obj = wb_obj.active

    row = sheet_obj.max_row
    column = sheet_obj.max_column

    print("Total Rows:", row)
    print("Total Columns:", column)

    values = []

    print("\nValue of first column")
    for i in range(row_start, row + row_start):
        cell_obj = sheet_obj.cell(row=i, column=column_no)
        print(cell_obj.value)
        values.append(cell_obj.value)


    values.remove(None)
    #FOR SQL

    findings_list = []

    forms_list = []

    step = 1

    completion = 0

    for x in values:

        hpercode = str(x)
        step = step + 1

        history = "SELECT COUNT(*) FROM hmrhisto where hpercode="+hpercode #NEED TO COUNT

        ward = "SELECT wardcode FROM hpatroom where hpercode ="+hpercode

        sas = "SELECT COUNT(*) AS SIGNS_AND_SYMPTOMS FROM hsignsymptoms WHERE enccode LIKE CONCAT('%',"+hpercode+", '%')" #NEED TO COUNT

        saspe = "SELECT COUNT(*) FROM hphyexam where hpercode="+hpercode #NEED TO COUNT

        course_in_the_ward = "select count(*) as COURSE_IN_THE_WARD from hcrsward where hpercode ="+hpercode #NEED TO COUNT

        diagnosis = "select tdcode, diagcode as DIAGNOSIS from hencdiag where tdcode = 'FINDX' and hpercode ="+hpercode

        doctor = '''select hencdiag.tdcode, hpersonal.lastname, hpersonal.firstname, hpersonal.middlename from hencdiag inner join hprovider ON hencdiag.licno=hprovider.licno

    inner join hpersonal ON hprovider.employeeid = hpersonal.employeeid where hencdiag.hpercode ='''+hpercode

        patient_name = "select patlast, patfirst, patmiddle from hperson where hpercode ="+hpercode

        print("Patient no: "+hpercode)

    #Query 1 History

        mycursor.execute(history)

        myresult = mycursor.fetchall()

        for x in myresult:

          

          if str(x[0]) == "0":

              findings_list.append("History: No History")

          else:

              print("History: "+str(x[0]))
              findings_list.append("History: "+str(x[0]))
              completion = completion + 1

    #Patient Query


          

        mycursor.execute(patient_name)

        myresult = mycursor.fetchall()

        for x in myresult:
            try:

                print("Patient name: "+' '.join(x))
                findings_list.append("Patient name: "+' '.join(x))
                

            except TypeError:

                print("Patient has no middle name")

                print("Patient name: "+str(x[0])+" "+str(x[1]))
                findings_list.append("Patient name: "+str(x[0])+" "+str(x[1]))


    #Ward query

        mycursor.execute(ward)

        myresult = mycursor.fetchall()

        for x in myresult:
          print("Ward: "+str(x[0]))

          if str(x[0]) == "0":

              findings_list.append("No Ward")

          else:
              
              findings_list.append("Ward: "+str(x[0]))
              

          


        #Query 2 SAS


        mycursor.execute(sas)

        myresult = mycursor.fetchall()

        for x in myresult:

          if str(x[0]) == "0":

              findings_list.append("Signs and symptoms: N/A")
          else:

              print("Signs and symptoms: "+str(x[0]))
              findings_list.append("Signs and symptoms: "+str(x[0]))
              completion = completion + 1
         


        #Query 3 SASPE


        mycursor.execute(saspe)

        myresult = mycursor.fetchall()

        for x in myresult:

          if str(x[0]) == "0":
              findings_list.append("PE: N/A")
          else:
              print("PE:"+str(x[0]))
              findings_list.append("PE:"+str(x[0]))
              completion = completion + 1
         


        #Query 4 COURSE IN THE WARD


        mycursor.execute(course_in_the_ward)

        myresult = mycursor.fetchall()

        for x in myresult:

            if str(x[0]) == "0":
              findings_list.append("Course in the ward: N/A")
            else:
              print("Course in the ward: "+str(x[0]))
              findings_list.append("Course in the ward: "+str(x[0]))
              completion = completion + 1
          


        #Query 5 DIAGNOSIS


        mycursor.execute(diagnosis)

        myresult = mycursor.fetchall()

        for x in myresult:

            if str(x[0]) == "FINDX" and str(x[1]) != None:

              completion = completion + 1
              findings_list.append("Diagnosis: "+str(x[0])+" , "+"ICD: "+" "+str(x[1]))

            else:

              print("Diagnosis: "+str(x[0])+" "+ str(x[1]))
              findings_list.append("Diagnosis: "+str(x[0])+" , "+"ICD: "+" "+str(x[1]))
        

          


       #Query 6 DOCTOR


        mycursor.execute(doctor)

        myresult = mycursor.fetchall()

        for x in myresult:
          print("Doctor: "+' '.join(x))
          findings_list.append("Doctor: "+' '.join(x))
          if x[0] == "FINDX":
              completion = completion + 1


          print("-------------")
          print("------END-------")

        if completion == 6:

            print("Record Complete")

        

        elif completion < 6:

            for forms in findings_list:
                input_value = sheet_obj[column_letter+str(step)]
                
            print(findings_list)
            input_value.value = '||'.join(findings_list)
            wb_obj.save(path+".xlsx")
            findings_list.clear()

            print("Completion:"+" "+str(completion))

        print("--END COMPLETION COUNT RESET--")

        completion = 0

       

        
            

    print(findings_list)

    if selected_dispo == "Admitted":

        sheet_obj.delete_cols(4)
        sheet_obj.delete_cols(4)
        sheet_obj.delete_cols(4)
        sheet_obj.delete_cols(4)
        sheet_obj.delete_cols(6)
        sheet_obj.delete_cols(8)
        sheet_obj.delete_cols(8)
        sheet_obj.delete_cols(8)
        sheet_obj.delete_cols(8)
        wb_obj.save(path+".xlsx")

    elif selected_dispo == "Discharged":

        cols = [1, 2, 3, 5, 5, 5, 5, 8]

        for x in cols:

            sheet_obj.delete_cols(x)
            wb_obj.save(path+".xlsx")
   
    label.configure(text="Report complete!")
   

#Initialize a Label to display the User Input
label=Label(win, text="", font=("Courier 22 bold"))
label.pack()

#Create an Entry widget to accept User Input
l1=Label(win, text="Enter filename", font=("Courier 12 bold"))
l1.pack()

path2= Entry(win, width= 40)
path2.focus_set()
path2.pack()

l5=Label(win, text="Enter Disposition report", font=("Courier 12 bold"))
l5.pack()

dispo = ttk.Combobox(win, values=["Admitted","Discharged"])
dispo.pack()

l2=Label(win, text="Enter column number of Health Record Number", font=("Courier 12 bold"))
l2.pack()

column_no2= Entry(win, width= 5)
column_no2.focus_set()
column_no2.pack()

l3=Label(win, text="Enter where row starts", font=("Courier 12 bold"))
l3.pack()

row_start2= Entry(win, width= 5)
row_start2.focus_set()
row_start2.pack()

l4=Label(win, text="Enter column letter", font=("Courier 12 bold"))
l4.pack()

column_letter2= Entry(win, width= 5)
column_letter2.focus_set()
column_letter2.pack()



#Create a Button to validate Entry Widget
ttk.Button(win, text= "Generate Result",width= 20, command= display_text).pack(pady=20)

win.mainloop()
